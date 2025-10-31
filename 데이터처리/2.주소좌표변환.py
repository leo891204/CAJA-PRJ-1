import pandas as pd
import requests
import re
from openpyxl import load_workbook
from openpyxl import Workbook


filepath = r'F:\OneDrive - 천안중앙고등학교\pythonworkspace\PROJECT\대학지도제작PRJ\고등교육기관 하반기 주소록(2021).xlsx'
excel_file = pd.read_excel(filepath,engine='openpyxl')
excel_file.columns = excel_file.loc[4].tolist()
excel_file = excel_file.drop(index=list(range(0,5)))


url = 'http://api.vworld.kr/req/address?'
main = 'service=address&request=getcoord&version=2.0&crs=epsg:4326&refine=true&simple=false&format=json&type='
road_address = 'ROAD'
road_address2= 'PARCEL'
address = '&address='
keys = '&key='
primary_key='C9D25EB1-7BCA-3FF3-B221-D8F0BF5A7427'

def request_geo(road):
    page = requests.get(url+main+road_address+address+road+keys+primary_key)
    json_data = page.json()
    if json_data['response']['status'] == 'OK':
        x=json_data['response']['result']['point']['x']
        y=json_data['response']['result']['point']['y']
        return x,y
    else:
        x=0
        y=0
        return x,y


try:
    wb = load_workbook(r"F:\OneDrive - 천안중앙고등학교\pythonworkspace\PROJECT\대학지도제작PRJ\고등교육기관 하반기 주소록(2021).xlsx", data_only=True)
    sheet = wb.active
except:
    wb=Workbook()
    sheet=wb.active

university_list = excel_file['학교명'].to_list()
address_list = excel_file['주소'].to_list()

for num,value in enumerate(address_list):
    addr = re.sub (r'\([^)]*\)','', value)
    print(addr)
    x,y = request_geo(addr)
    sheet.append([university_list[num], addr, x, y])
    
wb.save(r"F:\OneDrive - 천안중앙고등학교\pythonworkspace\PROJECT\대학지도제작PRJ\학교주소좌표전환결과.xlsx")
print('저장완료')